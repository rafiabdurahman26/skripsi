"""Laporan per sesi deteksi manusia + ringkasan Excel.

Mengikuti mekanisme reports project skripsi (drone_e99_face_recognition):
CSV per sesi + manifest JSON sebagai single source; Excel di-rebuild dari
source tiap kali (tidak pernah di-append).

Fokus laporan (tanpa koordinat bbox):
- track_id (traking IoU) -> jumlah subjek unik yang benar
- jumlah subjek terdeteksi, confidence score, ringkasan sesi

Struktur (prefix <ts> = YYYYMMDD_HHMMSS):
  reports/inference/session_<ts>/detection_log_<ts>.csv   (ts, frame, track_id, conf, fps)
  reports/inference/session_<ts>/per_second_<ts>.csv      (agregasi per detik)
  reports/inference/session_<ts>/subjects_<ts>.csv        (statistik per track/subjek)
  reports/inference/session_<ts>/session_stats_<ts>.csv   (ringkasan sesi)
  reports/inference/session_<ts>/persons_per_second_<ts>.png
  reports/runs/<ts>_detection.json                        (manifest)
  reports/summary/summary_detection.xlsx                  (rebuild tiap run)

Pemakaian di main.py:
    rep = SessionReporter(source, model, provider, conf)
    ... tiap frame: rep.add_frame(frame_idx, fps, dets)   # dets (m,5) x1,y1,x2,y2,conf
    rep.finalize(); export_summary()
"""
import csv
import glob
import json
import os
import time
from datetime import datetime

import numpy as np
import cv2
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.abspath(__file__))
REPORTS_DIR = os.path.join(ROOT, 'reports')
RUNS_DIR = os.path.join(REPORTS_DIR, 'runs')
SUMMARY_DIR = os.path.join(REPORTS_DIR, 'summary')

DETECTION_HEADERS = ['ts', 'frame_idx', 'track_id', 'conf', 'fps']
PER_SECOND_HEADERS = ['second', 'frame_count', 'persons', 'n_detections',
                      'mean_conf', 'max_conf', 'avg_fps']
SUBJECT_HEADERS = ['track_id', 'n_detections', 'mean_conf', 'max_conf',
                   'first_s', 'last_s', 'coverage_pct']
STATS_FIELDS = ['source', 'model', 'provider', 'conf_thres', 'duration_s',
                'frames', 'total_detections', 'unique_subjects',
                'max_concurrent', 'mean_persons_per_frame', 'mean_conf',
                'max_conf', 'avg_fps']
XLSX_DETECTION_HEADERS = ['timestamp', 'session_id', *STATS_FIELDS]
XLSX_SUBJECT_HEADERS = ['timestamp', 'session_id', *SUBJECT_HEADERS]

TRACK_IOU_THRESH = 0.3
TRACK_MAX_MISSED = 30   # ~1 detik @30fps sebelum track di-retire


def _now_ts():
    return datetime.now().strftime('%Y%m%d_%H%M%S')


def _iou(a, b):
    """IoU dua bbox (x1, y1, x2, y2)."""
    ix1, iy1 = max(a[0], b[0]), max(a[1], b[1])
    ix2, iy2 = min(a[2], b[2]), min(a[3], b[3])
    inter = max(0.0, ix2 - ix1) * max(0.0, iy2 - iy1)
    aa = (a[2] - a[0]) * (a[3] - a[1])
    bb = (b[2] - b[0]) * (b[3] - b[1])
    return inter / (aa + bb - inter + 1e-8)


class TrackTracker:
    """Pemberi track_id stabil agar jumlah subjek unik benar.

    Greedy match per deteksi (IoU terhadap track + pool retire), tanpa
    embedding. ponytail: orang keluar frame >1 dtk lalu kembali di posisi
    mirip -> id sama; di posisi baru -> id baru.
    """

    def __init__(self):
        self.next_id = 1
        self.tracks = {}    # id -> {'bbox': [x1,y1,x2,y2], 'missed': int}
        self.retired = []   # [(id, last_bbox)] pool re-ID posisi setelah miss
        self.stats = {}     # id -> {n, sum_conf, max_conf, first, last} (detik)

    def update(self, boxes, confs, sec):
        """Match deteksi satu frame -> list track_id sejajar boxes."""
        used = set()
        active = []
        for bbox, c in zip(boxes, confs):
            best, best_iou = None, 0.0
            for tid, t in self.tracks.items():
                if tid in used:
                    continue
                i = _iou(bbox, t['bbox'])
                if i > best_iou:
                    best_iou, best = i, tid
            if best is None or best_iou < TRACK_IOU_THRESH:
                best, best_iou = None, 0.0
                for tid, rbox in self.retired:
                    i = _iou(bbox, rbox)
                    if i > best_iou:
                        best_iou, best = i, tid
                if best is not None and best_iou >= TRACK_IOU_THRESH:
                    self.retired = [(i, b) for i, b in self.retired if i != best]
                else:
                    best = self.next_id
                    self.next_id += 1
            self.tracks[best] = {'bbox': [float(v) for v in bbox], 'missed': 0}
            used.add(best)
            st = self.stats.setdefault(best, {'n': 0, 'sum_conf': 0.0,
                                              'max_conf': 0.0,
                                              'first': sec, 'last': sec})
            st['n'] += 1
            st['sum_conf'] += float(c)
            st['max_conf'] = max(st['max_conf'], float(c))
            st['last'] = sec
            active.append(best)

        for tid in list(self.tracks):
            if tid in used:
                continue
            self.tracks[tid]['missed'] += 1
            if self.tracks[tid]['missed'] > TRACK_MAX_MISSED:
                self.retired.append((tid, self.tracks[tid]['bbox']))
                del self.tracks[tid]
        return active

    @property
    def n_unique_subjects(self):
        return len(self.stats)

    @property
    def n_detections(self):
        return sum(st['n'] for st in self.stats.values())


class SessionReporter:
    """Satu sesi live: add_frame() per frame, finalize() menulis laporan."""

    def __init__(self, source, model, provider, conf):
        self.source = source
        self.model = model
        self.provider = provider
        self.conf = conf
        self.ts = _now_ts()
        self.start = time.time()
        self.tracker = TrackTracker()
        self.per_sec = {}
        self.frames = 0
        self.sum_in_frame = 0.0
        self.max_in_frame = 0
        self.sum_conf = 0.0
        self.fps_sum = 0.0
        self.log_path = None
        self._csv = None
        self.session_stats = None

    def add_frame(self, frame_idx, fps, dets):
        """Satu frame diproses. dets: (m,5) x1,y1,x2,y2,conf (boleh kosong)."""
        sec = int(time.time() - self.start)
        rec = self.per_sec.setdefault(
            sec, {'frames': 0, 'ids': set(), 'n': 0, 'sc': 0.0, 'mx': 0.0, 'fs': 0.0})
        rec['frames'] += 1
        rec['fs'] += fps

        if self._csv is None:
            self._open_log()
        now = datetime.now()
        ts_str = f'{now.strftime("%H:%M:%S")}.{now.microsecond // 1000:03d}'
        n = len(dets)
        if n:
            ids = self.tracker.update(dets[:, :4].tolist(),
                                      dets[:, 4].tolist(), sec)
            for tid, c in zip(ids, dets[:, 4]):
                self._writer.writerow([ts_str, frame_idx, tid,
                                       f'{c:.4f}', f'{fps:.1f}'])
                rec['n'] += 1
                rec['sc'] += float(c)
                rec['mx'] = max(rec['mx'], float(c))
            rec['ids'].update(ids)

        self.frames += 1
        self.sum_in_frame += n
        self.max_in_frame = max(self.max_in_frame, n)
        self.fps_sum += fps

    def _open_log(self):
        session_dir = os.path.join(REPORTS_DIR, 'inference', f'session_{self.ts}')
        os.makedirs(session_dir, exist_ok=True)
        self.log_path = os.path.join(session_dir, f'detection_log_{self.ts}.csv')
        self._csv = open(self.log_path, 'w', newline='')
        self._writer = csv.writer(self._csv)
        self._writer.writerow(DETECTION_HEADERS)

    def finalize(self):
        """Tutup log; tulis per_second, subjects, stats, PNG, manifest."""
        if self._csv is not None:
            self._csv.close()
            self._csv = None
        if self.frames == 0:
            return False

        session_dir = os.path.join(REPORTS_DIR, 'inference', f'session_{self.ts}')

        per_second_path = os.path.join(session_dir, f'per_second_{self.ts}.csv')
        with open(per_second_path, 'w', newline='') as f:
            w = csv.writer(f)
            w.writerow(PER_SECOND_HEADERS)
            for sec in sorted(self.per_sec):
                r = self.per_sec[sec]
                w.writerow([sec, r['frames'], len(r['ids']), r['n'],
                            round(r['sc'] / r['n'], 4) if r['n'] else 0.0,
                            round(r['mx'], 4),
                            round(r['fs'] / r['frames'], 2) if r['frames'] else 0.0])

        subjects_path = os.path.join(session_dir, f'subjects_{self.ts}.csv')
        total_secs = max(self.per_sec) + 1
        with open(subjects_path, 'w', newline='') as f:
            w = csv.writer(f)
            w.writerow(SUBJECT_HEADERS)
            for tid in sorted(self.tracker.stats):
                st = self.tracker.stats[tid]
                w.writerow([tid, st['n'],
                            round(st['sum_conf'] / st['n'], 4) if st['n'] else 0.0,
                            round(st['max_conf'], 4), st['first'], st['last'],
                            round((st['last'] - st['first'] + 1) / total_secs * 100, 1)])

        dur = time.time() - self.start
        n_det = self.tracker.n_detections
        stats = {
            'source': self.source,
            'model': self.model,
            'provider': self.provider,
            'conf_thres': self.conf,
            'duration_s': round(dur, 2),
            'frames': self.frames,
            'total_detections': n_det,
            'unique_subjects': self.tracker.n_unique_subjects,
            'max_concurrent': self.max_in_frame,
            'mean_persons_per_frame': round(self.sum_in_frame / self.frames, 2),
            'mean_conf': round(self.sum_conf / n_det, 4) if n_det else 0.0,
            'max_conf': round(max(r['mx'] for r in self.per_sec.values()), 4),
            'avg_fps': round(self.fps_sum / self.frames, 2),
        }
        with open(os.path.join(session_dir, f'session_stats_{self.ts}.csv'),
                  'w', newline='') as f:
            w = csv.writer(f)
            w.writerow(STATS_FIELDS)
            w.writerow([stats[f] for f in STATS_FIELDS])

        self._plot_persons(os.path.join(session_dir,
                                        f'persons_per_second_{self.ts}.png'))
        self._write_manifest(stats)
        self.session_stats = stats
        print(f'[REPORT] sesi selesai: {dur:.2f}s, '
              f'{self.frames} frame, {n_det} deteksi, '
              f'{self.tracker.n_unique_subjects} subjek unik')
        return True

    def _plot_persons(self, path):
        """PNG: jumlah subjek unik per detik (oranye) vs FPS >= 20 (hijau)."""
        x = sorted(self.per_sec)
        y = [len(self.per_sec[s]['ids']) for s in x]
        y_fps = [self.per_sec[s]['fs'] / max(self.per_sec[s]['frames'], 1) for s in x]
        ymax = max(y) if y else 1
        H, W = 640, 1080
        bg = np.full((H, W, 3), 10, np.uint8)
        l, r, t, b = 70, W - 30, 40, H - 150
        scale_y = (b - t) / max(ymax, 1)
        for i, (v, fps) in enumerate(zip(y, y_fps)):
            xpix = l + (r - l) * i / max(len(x) - 1, 1)
            hh = max(int(v * scale_y), 1)
            cv2.rectangle(bg, (int(xpix), b - hh), (int(xpix + 8), b),
                          (0, 180, 255), -1)
            if fps >= 20:
                cv2.line(bg, (int(xpix), b), (int(xpix), b - 24), (60, 120, 60), 1)
        cv2.putText(bg, f'Subjek unik: {self.tracker.n_unique_subjects}',
                    (l, H - 100), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (255, 255, 255), 2)
        cv2.putText(bg, 'Bar oranye = orang per detik', (l, H - 55),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 200, 255), 1)
        cv2.putText(bg, 'Baris hijau = FPS >= 20', (500, H - 55),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.6, (120, 220, 120), 1)
        cv2.imwrite(path, bg)

    def _write_manifest(self, stats):
        os.makedirs(RUNS_DIR, exist_ok=True)
        manifest_path = os.path.join(RUNS_DIR, f'{self.ts}_manifest.json')
        with open(manifest_path, 'w') as f:
            json.dump({'timestamp': self.ts,
                       'file_name': f'{self.ts}_manifest.json',
                       'results': stats}, f, indent=2)


def export_summary():
    """Rebuild summary_detection.xlsx (Deteksi + Subjek) dari manifest & CSV."""
    os.makedirs(SUMMARY_DIR, exist_ok=True)
    manifests = sorted(glob.glob(os.path.join(RUNS_DIR, '*_manifest.json')),
                       reverse=True)
    det_rows, subj_rows = [], []
    for mpath in manifests:
        with open(mpath) as f:
            m = json.load(f)
        ts = m.get('timestamp', os.path.basename(mpath).split('_')[0])
        ts_cols = {'timestamp': ts, 'session_id': f'session_{ts}'}
        det_rows.append({**ts_cols, **m.get('results', {})})
        subj_csv = os.path.join(REPORTS_DIR, 'inference', f'session_{ts}',
                                f'subjects_{ts}.csv')
        if os.path.exists(subj_csv):
            for row in csv.DictReader(open(subj_csv, newline='')):
                subj_rows.append({**ts_cols, **row})
    det_rows = sorted(det_rows, key=lambda r: r['timestamp'])

    wb = Workbook()
    ws = wb.active or wb.create_sheet()
    ws.title = 'Deteksi'
    _xl_headers(ws, XLSX_DETECTION_HEADERS)
    for r in det_rows:
        ws.append([r[h] for h in XLSX_DETECTION_HEADERS])
    _autofit(ws)

    if subj_rows:
        ws2 = wb.create_sheet('Subjek')
        _xl_headers(ws2, XLSX_SUBJECT_HEADERS)
        for r in sorted(subj_rows, key=lambda r: r['timestamp']):
            ws2.append([r[h] for h in XLSX_SUBJECT_HEADERS])
        _autofit(ws2)

    wb.save(os.path.join(SUMMARY_DIR, 'summary_detection.xlsx'))
    print(f'[Report] summary_detection.xlsx OK ({len(det_rows)} sesi)')


def _xl_headers(ws, headers):
    fill = PatternFill('solid', fgColor='1F4E78')
    font = Font(bold=True, color='FFFFFF')
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = fill
        cell.font = font
    ws.freeze_panes = 'A2'


def _autofit(ws):
    for col in ws.columns:
        width = max(len(str(c.value)) for c in col if c.value is not None) + 2
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width, 60)


def _check():
    # ponytail: satu runnable check - tracker memberi id berulang stabil, dan
    #          reporter menulis laporan yang konsisten (2 detik simulasi)
    t = TrackTracker()
    ids1 = t.update([[0, 0, 50, 50]], [0.9], 0)
    ids2 = t.update([[3, 3, 53, 53]], [0.8], 0)   # geser 3px -> masih 1 orang
    assert ids1 == ids2 and t.n_unique_subjects == 1, 're-ID gagal untuk 1 orang'
    ids3 = t.update([[300, 300, 360, 360]], [0.6], 1)
    assert len(set(ids3)) == 1 and ids3[0] != ids2[0], 'orang baru harus id baru'
    assert t.n_unique_subjects == 2 and t.n_detections == 3
    print('[CHECK] tracker OK')

    r = SessionReporter('tello', 'm.onnx', 'CUDAExecutionProvider', 0.25)
    r.add_frame(0, 30.0, np.array([[0, 0, 50, 50, 0.9]], np.float32))
    r.add_frame(1, 30.0, np.array([[3, 3, 53, 53, 0.8]], np.float32))
    r.add_frame(2, 30.0, np.array([], np.float32).reshape(0, 5))
    assert r.finalize() and r.session_stats is not None
    assert r.session_stats['unique_subjects'] == 1
    assert r.session_stats['total_detections'] == 2
    per_sec = csv.DictReader(open(
        os.path.join(REPORTS_DIR, 'inference', f'session_{r.ts}',
                     f'per_second_{r.ts}.csv'), newline=''))
    rows = list(per_sec)
    assert rows and rows[0]['persons'] == '1', f'persons per detik salah: {rows}'
    manifest = glob.glob(os.path.join(RUNS_DIR, f'{r.ts}_manifest.json'))
    assert manifest, 'manifest tidak ditulis'
    export_summary()
    assert os.path.exists(os.path.join(SUMMARY_DIR, 'summary_detection.xlsx'))
    print('[CHECK] reporter OK')


if __name__ == '__main__':
    _check()
    print('[CHECK] reporting module import OK')